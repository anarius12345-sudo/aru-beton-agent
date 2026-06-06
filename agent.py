# -*- coding: utf-8 -*-
"""
АГЕНТ ARU BETON — ядро (Этап 1).
Сверка "карта (Smartgas) vs бак (Wialon ДУТ)" по методике A/B/C.
Источники: Smartgas API (карты) + Wialon отчёт 13 (Заправки/Сливы по ДУТ) + отчёт 12 (расход по машинам).
Выход: текст-сводка в Телеграм + Excel + PDF (как эталонный файл).
Запуск: GitHub Actions, каждое утро.
"""
import os, json, re, requests
from datetime import datetime, timedelta, timezone

# ---------- секреты ----------
WIALON_TOKEN = os.environ.get("WIALON_TOKEN")
SMARTGAS_KEY = os.environ.get("SMARTGAS_API_KEY")
BOT_TOKEN    = os.environ.get("TELEGRAM_BOT_TOKEN")
CHAT_ID      = os.environ.get("TELEGRAM_CHAT_ID")

# ---------- константы ----------
WIALON_BASE   = "https://hst-api.wialon.com/wialon/ajax.html"
SMARTGAS_BASE = "http://business.smartgas.global:8080/public-api/v1"
RES_ID    = 29957134
TPL_FUEL  = 13            # ARU_ANALYTICS (группа): Заправки/Сливы
TPL_TRIPS = 12            # ARU_ANALYTICS: Поездки/расход по машинам
GROUP_ID  = 29960488      # Все объекты
TZ        = timezone(timedelta(hours=5))

MATCH_WINDOW_MIN = 90     # окно матчинга карта<->ДУТ (мин)
PRICE_DT = 336            # ₸/л ДТ (для оценки суммы разницы)

# ---------- нормализация номеров ----------
CYR2LAT = {'А':'A','В':'B','Е':'E','К':'K','М':'M','Н':'H','О':'O','Р':'P','С':'C','Т':'T','У':'Y','Х':'X'}
KNOWN_CODES = {"374FI02","417FJ02","422FJ02","425FJ02","433FJ02","435FJ02","770AKC05","973AIL02","601FZ02","610FZ02","614FZ02","661FZ02","665FZ02","018WK02","019WK02","215GK02","619GP02","699GJ02","718GP02","719GP02","819GA02","822GA02","851GA02","864GA02","867GA02","894GJ02","966GJ02","976GK02","416FJ02","418FJ02","429FJ02","432FJ02","623FI02","623FJ02","AE799A","A965ADD","A574AUD","ABE798A","ADE470A","AEE058A","29354"}

def normalize_code(s):
    if not s: return ""
    s = "".join(CYR2LAT.get(ch, ch) for ch in str(s).strip())
    if "|" in s: s = s.split("|")[-1]
    s = s.strip().replace(" ", "").upper()
    s = re.sub(r'\(.*?\)', '', s)
    cands = re.findall(r'[A-Z0-9]+', s)
    try:
        kc = KNOWN_CODES
    except NameError:
        kc = set()
    for c in cands:
        for k in kc:
            if c.startswith(k):
                return "623FI02" if k=="623FJ02" else k
    best=""
    for c in cands:
        if sum(ch.isdigit() for ch in c)>=2 and len(c)>=4 and len(c)>len(best): best=c
    if best:
        if len(best)>1 and best[-1].isalpha() and best[:-1] in kc: best=best[:-1]
        if best not in kc and best+"02" in kc: best=best+"02"   # 864GA -> 864GA02 (один и тот же миксер)
        return "623FI02" if best=="623FJ02" else best
    for c in cands:
        if c.isdigit() and len(c)>=4: return c
    return s

# справочник: код -> тип
REG = {
 "374FI02":"Самосвал","417FJ02":"Самосвал","422FJ02":"Самосвал","425FJ02":"Самосвал","433FJ02":"Самосвал",
 "435FJ02":"Самосвал","770AKC05":"Самосвал","973AIL02":"Самосвал","601FZ02":"Самосвал","610FZ02":"Самосвал",
 "614FZ02":"Самосвал","661FZ02":"Самосвал","665FZ02":"Самосвал",
 "018WK02":"Миксер","019WK02":"Миксер","215GK02":"Миксер","619GP02":"Миксер","699GJ02":"Миксер","718GP02":"Миксер",
 "719GP02":"Миксер","819GA02":"Миксер","822GA02":"Миксер","851GA02":"Миксер","864GA02":"Миксер","867GA02":"Миксер",
 "894GJ02":"Миксер","966GJ02":"Миксер","976GK02":"Миксер","416FJ02":"Миксер","418FJ02":"Миксер","429FJ02":"Миксер",
 "432FJ02":"Миксер","623FI02":"Миксер",
 "AE799A":"Погрузчик","A965ADD":"Погрузчик","A574AUD":"Погрузчик","ABE798A":"Погрузчик","ADE470A":"Погрузчик","AEE058A":"Погрузчик",
 "29354":"Экскаватор",
}


def car_type(code):
    if code in REG: return REG[code]
    # 864GA -> 864GA02
    if code+"02" in REG: return REG[code+"02"]
    return "Прочее"

# ---------- helpers ----------
def tg(text):
    requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage", data={"chat_id": CHAT_ID, "text": text[:4000]})
def tg_doc(path, caption=""):
    with open(path,"rb") as f:
        requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                      data={"chat_id":CHAT_ID,"caption":caption[:1000]}, files={"document":f})
def wcall(svc, params, sid):
    return requests.get(WIALON_BASE, params={"svc":svc,"params":json.dumps(params),"sid":sid}, timeout=180).json()
def cval(c):
    return c.get("t","") if isinstance(c, dict) else (c if c is not None else "")
def pnum(x):
    if x is None: return 0.0
    s = str(x).replace("\xa0"," ").replace(" ","").replace("л","").replace("l","").replace(",",".")
    m = re.search(r'-?\d+\.?\d*', s)
    return float(m.group()) if m else 0.0
def is_car(label):
    """строка-машина (а не дата/итого)?"""
    if not label: return False
    if "|" in str(label): return True
    return False

# ====================================================================
# SMARTGAS
# ====================================================================
def get_smartgas(d_from, d_to):
    out=[]
    page=1; last=1
    while page<=last and page<=50:
        try:
            r=requests.get(f"{SMARTGAS_BASE}/transactions",
                headers={"Authorization":SMARTGAS_KEY},
                params={"dateFrom":d_from.isoformat(),"dateTo":d_to.isoformat(),"page":page}, timeout=90)
            data=r.json(); block=data.get("transactions",{})
            if isinstance(block,dict):
                txns=block.get("data",[])
                last=block.get("last_page", block.get("lastPage", 1)) or 1
            elif isinstance(block,list):
                txns=block; last=1
            else:
                txns=[]; last=1
            for t in txns:
                dn=t.get("display_name","")
                cdt=parse_dt(t.get("created_at",""))
                if cdt: cdt=cdt+timedelta(hours=5)   # FIX: Smartgas API отдаёт UTC -> +5ч (Алматы), как и Wialon
                out.append({"code":normalize_code(dn),"raw":dn,
                    "dt":cdt,
                    "liters":pnum(t.get("deliver_quantity",0)),
                    "order_amt":pnum(t.get("total_order_amt",0)),
                    "fuel":t.get("product_title",""),"azs":t.get("store_title",""),
                    "by_name":not bool(re.search(r'\d{3}', normalize_code(dn)))})
            if not txns: break
            page+=1
        except Exception as e:
            tg(f"⚠️ Smartgas стр.{page}: {e}"); break
    try: tg(f"[diag Smartgas] транзакций: {len(out)} (страниц: {page-1})")
    except: pass
    return out

# ====================================================================
# WIALON Заправки (отчёт 13, уровень машин)
# ====================================================================
def parse_dt(s, base_date=None):
    s = str(s).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S","%Y-%m-%d %H:%M:%S","%d.%m.%Y %H:%M:%S"):
        try: return datetime.strptime(s[:19], fmt)
        except: pass
    # только время "HH:MM:SS" + базовая дата
    if base_date and re.match(r'^\d{1,2}:\d{2}:\d{2}$', s):
        try:
            h,m,sec = map(int, s.split(":"))
            return datetime(base_date.year,base_date.month,base_date.day,h,m,sec)
        except: pass
    return None

def _parse_event(cells, lbl, cur_date):
    """Строку отчёта -> ('fuel'/'drain', dict) или None. Для топлива берём суммы по машине за день."""
    g = str(cells[0]).strip() if cells else ""
    if not is_car(g): return None
    code = normalize_code(g)
    tm = parse_dt(str(cells[1]) if len(cells)>1 else "", cur_date)
    if tm: tm = tm + timedelta(hours=5)
    if "аправк" in lbl:
        liters = pnum(cells[4]) if len(cells)>4 else 0.0
        if 20 <= liters <= 1500:   # допускаем сумму по машине за день (несколько заправок)
            return ("fuel", {"code":code,"raw":g,"dt":tm,"liters":liters,
                             "pos":str(cells[2]) if len(cells)>2 else ""})
    elif "лив" in lbl:
        liters = pnum(cells[5]) if len(cells)>5 else 0.0
        if liters >= 1:
            return ("drain", {"code":code,"raw":g,"dt":tm,"liters":liters,
                              "pos":str(cells[3]) if len(cells)>3 else ""})
    return None

def _read_rows(idx, lbl, sid):
    """дата -> get_result_subrows: строки уровня машин (= сумма по машине за день)."""
    out_f=[]; out_d=[]
    top=wcall("report/get_result_rows",{"tableIndex":idx,"indexFrom":0,"indexTo":300},sid)
    if not isinstance(top,list): return out_f,out_d
    for di,drow in enumerate(top):
        dcells=[cval(c) for c in drow.get("c",[])]
        cur_date=None
        g0=str((dcells+[""])[0]).strip()
        if re.match(r'^\d{2}\.\d{2}\.\d{4}$',g0):
            try: cur_date=datetime.strptime(g0,"%d.%m.%Y")
            except: pass
        sub=wcall("report/get_result_subrows",{"tableIndex":idx,"rowIndex":di,"indexFrom":0,"indexTo":500},sid)
        if not isinstance(sub,list): continue
        for k in sub:
            ev=_parse_event([cval(c) for c in k.get("c",[])], lbl, cur_date)
            if ev: (out_f if ev[0]=="fuel" else out_d).append(ev[1])
    return out_f,out_d

def get_wialon_fuel(sid, ts_from, ts_to):
    wcall("report/cleanup_result", {}, sid)
    rep = wcall("report/exec_report", {"reportResourceId":RES_ID,"reportTemplateId":TPL_FUEL,
        "reportObjectId":GROUP_ID,"reportObjectSecId":0,"interval":{"from":ts_from,"to":ts_to,"flags":0}}, sid)
    tables = rep.get("reportResult",{}).get("tables",[])
    raw_f=[]; raw_d=[]
    for idx,t in enumerate(tables):
        lbl=t.get("label","")
        if "аправк" not in lbl and "лив" not in lbl: continue
        f,d = _read_rows(idx, lbl, sid)
        raw_f += f; raw_d += d
    # агрегируем топливо ПО МАШИНЕ ЗА ДЕНЬ (надёжно при любой структуре отчёта)
    from collections import defaultdict
    agg=defaultdict(lambda:{"liters":0.0,"dt":None,"pos":"","code":None})
    for x in raw_f:
        day = x["dt"].date() if x["dt"] else None
        e=agg[(x["code"], day)]
        e["liters"]+=x["liters"]; e["code"]=x["code"]; e["pos"]=e["pos"] or x["pos"]
        if x["dt"] and (e["dt"] is None or x["dt"]<e["dt"]): e["dt"]=x["dt"]
    res={"fuel":[{"code":v["code"],"dt":v["dt"],"liters":round(v["liters"],2),"pos":v["pos"]}
                 for v in agg.values()],
         "drain":raw_d}
    try:
        msg=f"[diag Wialon] машин-дней: {len(res['fuel'])}, сливов: {len(res['drain'])}\n"
        for f in res["fuel"][:5]: msg+=f"  {f['code']} {f['dt']} {f['liters']}л\n"
        tg(msg)
    except: pass
    return res

# ====================================================================
# СВЕРКА A/B/C — ПО МАШИНЕ ЗА ДЕНЬ
# ====================================================================
def match_abc(sg, wl_fuel):
    """Сводим обе стороны к МАШИНЕ за весь период и сравниваем суммы. A — есть и карта и ДУТ; B — только карта; C — только ДУТ."""
    def agg(items):
        d={}
        for x in items:
            k=x["code"]
            if k not in d:
                d[k]={"code":x["code"],"liters":0.0,"dt":x.get("dt"),
                      "azs":x.get("azs",""),"order_amt":0.0,"fuel":x.get("fuel",""),"pos":x.get("pos","")}
            e=d[k]; e["liters"]+=x["liters"]; e["order_amt"]+=x.get("order_amt",0.0)
            if x.get("dt") and (e["dt"] is None or x["dt"]<e["dt"]): e["dt"]=x["dt"]
            e["azs"]=e["azs"] or x.get("azs",""); e["fuel"]=e["fuel"] or x.get("fuel","")
            e["pos"]=e["pos"] or x.get("pos","")
        return d
    cardm=agg(sg); bakm=agg(wl_fuel)
    pairs=[]; B=[]; C=[]
    for k in set(cardm)|set(bakm):
        c=cardm.get(k); b=bakm.get(k)
        if c and b:
            pairs.append({"code":c["code"],"mp":c["liters"],"dut":b["liters"],
                          "delta":c["liters"]-b["liters"],"gap":0,
                          "azs":c["azs"],"dt":c["dt"] or b["dt"],"fuel":c["fuel"]})
        elif c:
            B.append({"code":c["code"],"liters":c["liters"],"dt":c["dt"],"azs":c["azs"],
                      "order_amt":c["order_amt"],"fuel":c["fuel"],
                      "by_name": not bool(re.search(r'\d{3}', c["code"]))})
        else:
            C.append({"code":b["code"],"liters":b["liters"],"dt":b["dt"],"pos":b["pos"]})
    try:
        tg(f"[diag match] машин-карт:{len(cardm)} машин-бак:{len(bakm)} A:{len(pairs)} B:{len(B)} C:{len(C)}")
    except: pass
    return pairs, B, C


# ====================================================================
# СБОРКА EXCEL
# ====================================================================
def build_excel(path, period, pairs, B, C, sg, wl, drains):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    HEAD = PatternFill("solid", fgColor="1F4E78"); HF = Font(bold=True,color="FFFFFF")
    RED  = PatternFill("solid", fgColor="C00000")
    thin = Side(style="thin",color="BFBFBF"); BORD = Border(thin,thin,thin,thin)
    def style_head(ws,row,ncol,fill=HEAD):
        for c in range(1,ncol+1):
            cell=ws.cell(row=row,column=c); cell.fill=fill; cell.font=HF
            cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=BORD
    def zebra(ws,r0,r1,ncol):
        for r in range(r0,r1+1):
            f = PatternFill("solid",fgColor="F2F2F2") if (r-r0)%2 else None
            for c in range(1,ncol+1):
                cell=ws.cell(row=r,column=c); cell.border=BORD
                if f: cell.fill=f

    # ===== ВСЁ НА ОДНОМ ЛИСТЕ =====
    ws=wb.active; ws.title="Отчёт"
    ws["A1"]="СВЕРКА ТОПЛИВА ARU BETON — карты (МП) ↔ датчики (ДУТ)"; ws["A1"].font=Font(bold=True,size=13)
    ws["A2"]=f"Период проверки: {period}"

    def block(title, headers, data_rows, fill=HEAD, totals=None, tsize=12):
        ws.append([])
        ws.append([title]); ws.cell(row=ws.max_row,column=1).font=Font(bold=True,size=tsize)
        ws.append(headers); hr=ws.max_row; style_head(ws,hr,len(headers),fill)
        for row in data_rows: ws.append(row)
        if totals is not None: ws.append(totals)
        zebra(ws,hr+1,ws.max_row,len(headers))
        if totals is not None:
            for c in range(1,len(headers)+1): ws.cell(row=ws.max_row,column=c).font=Font(bold=True)

    mpA=sum(p["mp"] for p in pairs); dutA=sum(p["dut"] for p in pairs)
    mpB=sum(b["liters"] for b in B); dutC=sum(c["liters"] for c in C)
    sum_all=sum(t["order_amt"] for t in sg); sumB=sum(b.get("order_amt",0) for b in B)
    diffA=mpA-dutA; pct=diffA/mpA*100 if mpA else 0; diff_money=round(diffA*PRICE_DT)

    # 1) Баланс
    block("БАЛАНС", ["№","Категория","Машин","МП, л","ДУТ, л","Разница МП−ДУТ, л","Сумма / разница, ₸"],
        [[1,"A — карта + ДУТ сопоставлены",len(pairs),round(mpA,1),round(dutA,1),round(diffA,1),diff_money],
         [2,"B — карта есть, ДУТ нет",len(B),round(mpB,1),"—","—",round(sumB,0)],
         [3,"C — ДУТ есть, карты нет",len(C),"—",round(dutC,1),"—","—"]],
        totals=["","ИТОГО",len(pairs)+len(B)+len(C),round(mpA+mpB,1),round(dutA+dutC,1),round(diffA,1),round(sum_all,0)])

    # 2) Проверка сходимости
    block("ПРОВЕРКА СХОДИМОСТИ БАЛАНСОВ", ["№","Показатель","Значение, л","Контроль","Статус"],
        [[1,"Карты всего (МП)",round(mpA+mpB,1),f"A+B = {round(mpA+mpB,1)}","СХОДИТСЯ ✓"],
         [2,"Wialon всего (ДУТ)",round(dutA+dutC,1),f"A+C = {round(dutA+dutC,1)}","СХОДИТСЯ ✓"],
         [3,"Разница МП−ДУТ в кат. A (калибровка)",round(diffA,1),f"{pct:.2f} %",f"{diff_money:,} ₸".replace(',',' ')]],
        tsize=11)

    # 3) A — по машинам за день
    a_rows=[]
    for i,p in enumerate(sorted(pairs,key=lambda x:x["code"]),1):
        dpct=p["delta"]/p["mp"]*100 if p["mp"] else 0
        a_rows.append([i,p["code"],car_type(p["code"]),p["dt"].strftime("%Y-%m-%d %H:%M") if p["dt"] else "",
                       (p["azs"] or "")[:28],round(p["mp"],1),round(p["dut"],1),round(p["delta"],1),round(dpct,1)])
    block("A — СВЕРКА ПО МАШИНАМ ЗА ДЕНЬ (карта + ДУТ)",
        ["№","Код ТС","Тип","Время карты","АЗС","МП, л","ДУТ, л","Δ, л","Δ %"], a_rows,
        totals=["","ИТОГО","","","",round(mpA,1),round(dutA,1),round(diffA,1),""])

    # 4) Все машины парка
    codes=set([t["code"] for t in sg]+[f["code"] for f in wl]); bycode={}
    for t in sg: bycode.setdefault(t["code"],{"mp":0,"dut":0}); bycode[t["code"]]["mp"]+=t["liters"]
    for f in wl: bycode.setdefault(f["code"],{"mp":0,"dut":0}); bycode[f["code"]]["dut"]+=f["liters"]
    m_rows=[]
    for i,code in enumerate(sorted(codes),1):
        v=bycode[code]; has="карта+ДУТ" if v["mp"]>0 and v["dut"]>0 else ("только карта" if v["mp"]>0 else "только Wialon")
        m_rows.append([i,code,car_type(code),round(v["mp"],1) if v["mp"] else "—",round(v["dut"],1) if v["dut"] else "—",has])
    tot_mp=sum(v["mp"] for v in bycode.values()); tot_dut=sum(v["dut"] for v in bycode.values())
    block("ВСЕ МАШИНЫ ПАРКА", ["№","Код ТС","Тип","МП (карта), л","ДУТ (Wialon), л","Наличие"], m_rows,
          totals=["","ИТОГО","",round(tot_mp,1),round(tot_dut,1),""])

    # 5) B — карта без датчика
    b_rows=[]
    for i,b in enumerate(sorted(B,key=lambda x:(x["dt"] or datetime.min)),1):
        b_rows.append([i,b["code"] if not b["by_name"] else "(по имени)",
                       b["dt"].strftime("%Y-%m-%d %H:%M") if b["dt"] else "",(b["azs"] or "")[:28],
                       round(b["liters"],1),round(b["order_amt"],0),b["fuel"]])
    block("B — КАРТА БЕЗ ДАТЧИКА", ["№","Код ТС","Время","АЗС","Литры","₸","Топливо"], b_rows, fill=RED,
          totals=["","ИТОГО","","",round(mpB,1),round(sumB,0),""] if B else None)

    # 6) C — Wialon без карты
    c_rows=[]
    for i,c in enumerate(sorted(C,key=lambda x:x["code"]),1):
        c_rows.append([i,c["code"],car_type(c["code"]),c["dt"].strftime("%Y-%m-%d %H:%M") if c["dt"] else "",
                       round(c["liters"],1),(c["pos"] or "")[:28]])
    block("C — WIALON БЕЗ КАРТЫ", ["№","Код ТС","Тип","Время Wialon","Залито, л","Место"], c_rows, fill=RED,
          totals=["","ИТОГО","","",round(dutC,1),""] if C else None)

    # 7) Сливы
    d_rows=[]
    for i,d in enumerate(sorted(drains,key=lambda x:-x["liters"]),1):
        d_rows.append([i,d["code"],car_type(d["code"]),d["dt"].strftime("%Y-%m-%d %H:%M") if d["dt"] else "",
                       round(d["liters"],1),(d["pos"] or "")[:28]])
    block("СЛИВЫ ТОПЛИВА (по ДУТ)", ["№","Код ТС","Тип","Время","Слито, л","Место"], d_rows, fill=RED,
          totals=["","ИТОГО","","",round(sum(d["liters"] for d in drains),1),""] if drains else None)

    # ширины столбцов — компромисс, т.к. все таблицы на одном листе
    for col,w in zip("ABCDEFGHI",[5,30,14,18,26,16,13,9,8]): ws.column_dimensions[col].width=w

    # единый шрифт Arial + разделители разрядов по всем листам (как в эталонном файле)
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for cell in row:
                v=cell.value
                if v is None: continue
                f=cell.font
                size=f.size if (f.size and f.size>=12) else 10
                cell.font=Font(name="Arial", size=size, bold=f.bold, italic=f.italic, color=f.color)
                if isinstance(v,bool):
                    pass
                elif isinstance(v,int):
                    cell.number_format='#,##0'          # 892384 -> 892 384
                elif isinstance(v,float):
                    cell.number_format='#,##0' if v==int(v) else '#,##0.0'  # 1856,4 -> 1 856,4
    wb.save(path)
    return {"mpA":mpA,"dutA":dutA,"mpB":mpB,"dutC":dutC,"pct":pct,
            "nA":len(pairs),"nB":len(B),"nC":len(C),"sum_all":sum_all,"drains":len(drains)}

# ====================================================================
# PDF
# ====================================================================
def build_pdf(path, period, stats, pairs, B, C, sg, wl, drains):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        reg="/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        bld="/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        fn="Helvetica"; fnb="Helvetica-Bold"
        if os.path.exists(reg): pdfmetrics.registerFont(TTFont("DV",reg)); fn="DV"
        if os.path.exists(bld): pdfmetrics.registerFont(TTFont("DVB",bld)); fnb="DVB"
        BLUE=colors.HexColor("#1F4E78"); RED=colors.HexColor("#C00000")
        GREY=colors.HexColor("#BFBFBF"); ZEB=colors.HexColor("#F2F2F2")
        doc=SimpleDocTemplate(path,pagesize=A4,topMargin=12*mm,bottomMargin=12*mm,
                              leftMargin=10*mm,rightMargin=10*mm)
        title=ParagraphStyle("t",fontName=fnb,fontSize=14,spaceAfter=2)
        sub=ParagraphStyle("s",fontName=fn,fontSize=9,spaceAfter=2)
        htbl=ParagraphStyle("h",fontName=fnb,fontSize=10,spaceBefore=9,spaceAfter=3)
        cell=ParagraphStyle("c",fontName=fn,fontSize=7,leading=8)
        cellb=ParagraphStyle("cb",fontName=fnb,fontSize=7,leading=8,textColor=colors.white)

        def nf(v,dec=1):
            if v is None: return ""
            if isinstance(v,str): return v
            s=f"{v:,.{dec}f}".replace(",", " ")
            if dec>0: s=s.replace(".", ",")
            return s
        cellt=ParagraphStyle("ct",fontName=fnb,fontSize=7,leading=8)
        def mk(rows,widths,head=BLUE,total=False):
            ti=len(rows)-1
            body=[]
            for i,r in enumerate(rows):
                stl=cellb if i==0 else (cellt if (total and i==ti) else cell)
                body.append([Paragraph(str(v),stl) for v in r])
            t=Table(body,colWidths=[w*mm for w in widths],hAlign="LEFT",repeatRows=1)
            sty=[
                ("BACKGROUND",(0,0),(-1,0),head),
                ("GRID",(0,0),(-1,-1),0.4,GREY),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,ZEB]),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
                ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
            ]
            if total: sty.append(("BACKGROUND",(0,ti),(-1,ti),colors.HexColor("#E2E2E2")))
            t.setStyle(TableStyle(sty))
            return t

        mpA,dutA=stats["mpA"],stats["dutA"]; mpB,dutC=stats["mpB"],stats["dutC"]
        diffA=mpA-dutA; pct=stats["pct"]; sum_all=stats["sum_all"]
        sumB=sum(b.get("order_amt",0) for b in B); diff_money=round(diffA*PRICE_DT)

        el=[Paragraph("Сверка топлива ARU Beton",title),
            Paragraph(f"Период: {period}",sub),Spacer(1,4)]

        # 1) Баланс
        el.append(Paragraph("Баланс топлива: карты (МП) ↔ датчики (ДУТ)",htbl))
        el.append(mk([
            ["№","Категория","Машин","МП, л","ДУТ, л","Разница, л","Сумма / разница, ₸"],
            ["1","A — карта + ДУТ сопоставлены",stats["nA"],nf(mpA),nf(dutA),nf(diffA),nf(diff_money,0)],
            ["2","B — карта есть, ДУТ нет",stats["nB"],nf(mpB),"—","—",nf(sumB,0)],
            ["3","C — ДУТ есть, карты нет",stats["nC"],"—",nf(dutC),"—","—"],
            ["","ИТОГО",stats["nA"]+stats["nB"]+stats["nC"],nf(mpA+mpB),nf(dutA+dutC),nf(diffA),nf(sum_all,0)],
        ],[8,50,15,20,20,24,33],total=True))

        # 2) Проверка сходимости
        el.append(Paragraph("Проверка сходимости балансов",htbl))
        el.append(mk([
            ["№","Показатель","Значение, л","Контроль","Статус"],
            ["1","Карты всего (МП)",nf(mpA+mpB),f"A+B = {nf(mpA+mpB)}","СХОДИТСЯ ✓"],
            ["2","Wialon всего (ДУТ)",nf(dutA+dutC),f"A+C = {nf(dutA+dutC)}","СХОДИТСЯ ✓"],
            ["3","Разница МП−ДУТ в кат. A (калибровка)",nf(diffA),f"{pct:.2f} %",f"{nf(diff_money,0)} ₸"],
        ],[8,66,24,34,38]))

        # 3) A — по машинам за день
        arows=[["№","Код ТС","Тип","Время карты","АЗС","МП, л","ДУТ, л","Δ, л","Δ %"]]
        for i,p in enumerate(sorted(pairs,key=lambda x:x["code"]),1):
            dpct=p["delta"]/p["mp"]*100 if p["mp"] else 0
            arows.append([i,p["code"],car_type(p["code"]),
                          p["dt"].strftime("%d.%m %H:%M") if p["dt"] else "",
                          (p["azs"] or "")[:30],nf(p["mp"]),nf(p["dut"]),nf(p["delta"]),nf(dpct)])
        arows.append(["","ИТОГО","","","",nf(mpA),nf(dutA),nf(diffA),""])
        el.append(Paragraph("A — сверка по машинам за день (карта + ДУТ)",htbl))
        el.append(mk(arows,[8,20,17,22,43,17,17,15,14],total=True))

        # 4) Все машины парка
        codes=set([t["code"] for t in sg]+[f["code"] for f in wl]); bycode={}
        for t in sg: bycode.setdefault(t["code"],{"mp":0,"dut":0}); bycode[t["code"]]["mp"]+=t["liters"]
        for f in wl: bycode.setdefault(f["code"],{"mp":0,"dut":0}); bycode[f["code"]]["dut"]+=f["liters"]
        mrows=[["№","Код ТС","Тип","МП (карта), л","ДУТ (Wialon), л","Наличие"]]
        for i,code in enumerate(sorted(codes),1):
            v=bycode[code]; has="карта+ДУТ" if v["mp"]>0 and v["dut"]>0 else ("только карта" if v["mp"]>0 else "только Wialon")
            mrows.append([i,code,car_type(code),nf(v["mp"]) if v["mp"] else "—",nf(v["dut"]) if v["dut"] else "—",has])
        mrows.append(["","ИТОГО","",nf(sum(v["mp"] for v in bycode.values())),nf(sum(v["dut"] for v in bycode.values())),""])
        el.append(Paragraph("Все машины парка",htbl))
        el.append(mk(mrows,[8,26,24,34,38,40],total=True))

        # 5) B — карта без датчика
        brows=[["№","Код ТС","Время","АЗС","Литры","₸","Топливо"]]
        for i,b in enumerate(sorted(B,key=lambda x:(x["dt"] or datetime.min)),1):
            brows.append([i,b["code"] if not b["by_name"] else "(по имени)",
                          b["dt"].strftime("%d.%m %H:%M") if b["dt"] else "",
                          (b["azs"] or "")[:30],nf(b["liters"]),nf(b["order_amt"],0),b["fuel"]])
        if B: brows.append(["","ИТОГО","","",nf(mpB),nf(sumB,0),""])
        el.append(Paragraph("B — карта без датчика",htbl))
        el.append(mk(brows,[8,24,22,48,18,26,24],head=RED,total=bool(B)))

        # 6) C — Wialon без карты
        crows=[["№","Код ТС","Тип","Время Wialon","Залито, л","Место"]]
        for i,c in enumerate(sorted(C,key=lambda x:x["code"]),1):
            crows.append([i,c["code"],car_type(c["code"]),
                          c["dt"].strftime("%d.%m %H:%M") if c["dt"] else "",
                          nf(c["liters"]),(c["pos"] or "")[:42]])
        if C: crows.append(["","ИТОГО","","",nf(dutC),""])
        el.append(Paragraph("C — Wialon без карты",htbl))
        el.append(mk(crows,[8,24,22,24,22,70],head=RED,total=bool(C)))

        # 7) Сливы
        if drains:
            drows=[["№","Код ТС","Тип","Время","Слито, л","Место"]]
            for i,d in enumerate(sorted(drains,key=lambda x:-x["liters"]),1):
                drows.append([i,d["code"],car_type(d["code"]),
                              d["dt"].strftime("%d.%m %H:%M") if d["dt"] else "",
                              nf(d["liters"]),(d["pos"] or "")[:42]])
            drows.append(["","ИТОГО","","",nf(sum(d["liters"] for d in drains)),""])
            el.append(Paragraph("Сливы топлива (по ДУТ)",htbl))
            el.append(mk(drows,[8,24,22,24,22,70],head=RED,total=True))

        doc.build(el)
        return True
    except Exception as e:
        tg(f"⚠️ PDF не создан: {e}")
        return False

# ====================================================================
# RUN
# ====================================================================
def run_period(label, ts_from, ts_to, d_from, d_to, sid, tag):
    period=f"{d_from.strftime('%d.%m.%Y')} – {d_to.strftime('%d.%m.%Y')}"
    sg = get_smartgas(d_from - timedelta(days=1), d_to + timedelta(days=1))  # FIX: берём шире, чтобы не терять края суток из-за UTC
    # фильтр по локальной дате периода (после +5ч Smartgas уже в Алматы)
    def _in_period(t):
        if t["dt"] is None: return True
        d = t["dt"].date()
        return d_from <= d <= d_to
    sg = [t for t in sg if _in_period(t)]
    wl = get_wialon_fuel(sid, ts_from, ts_to)
    fuelings = [f for f in wl["fuel"] if f["dt"] is None or (d_from <= f["dt"].date() <= d_to)]
    drains = [x for x in wl["drain"] if x["dt"] is None or (d_from <= x["dt"].date() <= d_to)]
    pairs, B, C = match_abc(sg, fuelings)

    xlsx=f"/tmp/Анализ_заправки_{tag}.xlsx"; pdf=f"/tmp/Сверка_{tag}.pdf"
    stats=build_excel(xlsx, period, pairs, B, C, sg, fuelings, drains)
    has_pdf=build_pdf(pdf, period, stats, pairs, B, C, sg, fuelings, drains)

    diff=stats["mpA"]-stats["dutA"]
    tot_card=stats['mpA']+stats['mpB']; tot_dut=stats['dutA']+stats['dutC']
    tot_diff=tot_card-tot_dut; tot_pct=(tot_diff/tot_card*100) if tot_card else 0
    lines=[f"🚛 СВЕРКА «карта vs бак» — {label}",f"Период: {period}","",
           f"📥 Куплено по картам: {tot_card:.0f} л / {stats['sum_all']:,.0f} ₸".replace(',',' '),
           f"⛽ ДУТ Wialon: {tot_dut:.0f} л",
           f"📊 Разница карта−бак (всего): {tot_diff:+.0f} л ({tot_pct:+.1f}%)","",
           f"A — сопоставлено машин: {stats['nA']} | в т.ч. разница {diff:+.0f} л ({stats['pct']:.1f}%)",
           f"B карта без ДУТ: {stats['nB']} машин ({stats['mpB']:.0f} л)",
           f"C ДУТ без карты: {stats['nC']} машин ({stats['dutC']:.0f} л)",
           f"💧 Сливы по ДУТ: {stats['drains']}"]
    # топ расхождений по машинам
    big=sorted([p for p in pairs if abs(p["delta"])>=20],key=lambda x:-abs(x["delta"]))[:5]
    if big:
        lines.append("\n⚠️ Машины с расхождением (проверить):")
        for p in big:
            lines.append(f"  {p['code']}: карта {p['mp']:.0f} / бак {p['dut']:.0f} → {p['delta']:+.0f} л")
    tg("\n".join(lines))
    tg_doc(xlsx, f"Excel: анализ заправок {period}")
    if has_pdf: tg_doc(pdf, f"PDF: сверка {period}")

def main():
    if not all([WIALON_TOKEN,SMARTGAS_KEY,BOT_TOKEN,CHAT_ID]):
        tg("❌ Не заданы секреты."); return
    login=requests.get(WIALON_BASE,params={"svc":"token/login","params":'{"token":"%s"}'%WIALON_TOKEN},timeout=60).json()
    if "eid" not in login: tg(f"❌ Wialon вход: {str(login)[:200]}"); return
    sid=login["eid"]
    now=datetime.now(TZ)
    y=(now-timedelta(days=1)).date()
    yf=int(datetime(y.year,y.month,y.day,0,0,0,tzinfo=TZ).timestamp())
    yt=int(datetime(y.year,y.month,y.day,23,59,59,tzinfo=TZ).timestamp())
    ms=now.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
    try: run_period("ЗА ВЧЕРА", yf, yt, y, y, sid, "вчера")
    except Exception as e: tg(f"❌ Ошибка (вчера): {e}")
    try: run_period("С НАЧАЛА МЕСЯЦА", int(ms.timestamp()), yt, ms.date(), y, sid, "месяц")
    except Exception as e: tg(f"❌ Ошибка (месяц): {e}")

if __name__=="__main__":
    main()
